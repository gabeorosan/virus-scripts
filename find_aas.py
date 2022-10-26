import scipy
import scipy.spatial as spatial
import numpy as np
import pandas as pd
import sys

pdb_url = sys.argv[1]
pa_url = sys.argv[2]

with open(pdb_url) as pdb_file:
    pdb_coords = {}
    pdb_info = {}
    for l in pdb_file.readlines():
        chain = l[21]
        x = float(l[30:38])
        y = float(l[38:46])
        z = float(l[46:54])
        aa = l[17:20]
        rn = l[23:26]
        atom = l[13:16]
        if chain not in pdb_coords.keys():
            pdb_coords[chain] = []
        pdb_coords[chain].append([x, y, z])
        if chain not in pdb_info.keys():
            pdb_info[chain] = []
        pdb_info[chain].append({'aa_rn': aa + rn, 'atom': atom})
#read pa file
with open(pa_url) as pa_file:
    pa_coords = []
    for l in pa_file.readlines():
        x = float(l[30:38])
        y = float(l[38:46])
        z = float(l[46:54])
        pa_coords.append([x, y, z])

dfs = []
point_df = pd.DataFrame({'Point': range(1, len(pa_coords)+1)})
for c in pdb_coords.keys():
    distances = np.transpose(scipy.spatial.distance.cdist(pdb_coords[c], pa_coords))
    d_argsort = np.argsort(distances, axis=1)
    ds, atoms, aas, aa_others = [], [], [], []
    for p_n in range(len((d_argsort))):
        min_ix = d_argsort[p_n][0]
        d = distances[p_n][min_ix]
        aa_rn = pdb_info[c][min_ix]['aa_rn']
        atom = pdb_info[c][min_ix]['atom']

        for d_ix in d_argsort[p_n]:
            if distances[p_n][d_ix] > 5:
                aa_others.append('N/A')
                break
            else:
                other_aa_rn = pdb_info[c][d_ix]['aa_rn']
                if aa_rn != other_aa_rn:
                    aa_others.append(other_aa_rn)
                    break
        ds.append(d)
        atoms.append(atom)
        aas.append(aa_rn)


    df = pd.DataFrame({'Distance': ds, 'Atom': atoms, 'AA/RN': aas, 'Other AA': aa_others})
    dfs.append(df)

from openpyxl import load_workbook
from openpyxl.styles import Alignment

xl_file = pdb_url.split('/')[-1].split('.')[0] + '.xlsx'
pa_name=pa_url.split('/')[-1].split('.')[0]
try:
    with pd.ExcelWriter(xl_file, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
        point_df.to_excel(writer, startrow=1, sheet_name=pa_name, index = False)
except: 
    point_df.to_excel(xl_file, sheet_name=pa_name, startrow = 1, index = False)

def excel_style(col):
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    if not col: return 'A'
    repeat = False
    while col:
        col, rem = divmod(col, 26)
        if repeat: rem -= 1
        result[:0] = LETTERS[rem]
        repeat = True
    return ''.join(result)

for df in dfs:
    with pd.ExcelWriter(xl_file, engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:  
        book = load_workbook(xl_file)
        for sheetname in writer.sheets:
            if sheetname != pa_name: continue
            scol = writer.sheets[sheetname].max_column
            mstring = excel_style(scol) + '1' + ':' + excel_style(scol+3) + '1'
            writer.sheets[sheetname].merge_cells(mstring)
            cstring = excel_style(scol).upper() + '1'
            writer.sheets[sheetname][cstring].alignment = Alignment(horizontal='center')
            writer.sheets[sheetname][cstring] = 'Chain ' + excel_style(scol // 4).upper()
            df.to_excel(writer, startrow = 1, sheet_name=sheetname, startcol=scol, index = False)


